import openpyxl
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


#sheet selection
wb=load_workbook('pivot_tables.xlsx')
sheet=wb['spending_gender']

#selection of active rows

min_column=wb.active.min_column
max_column=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row

barchart=BarChart()

data=Reference(sheet,
          min_col=min_column+1,
          max_col=max_column,
          min_row=min_row,
          max_row=max_row)

catogeries=Reference(sheet,
          min_col=min_column,
          max_col=min_column,
          min_row=min_row+1,
          max_row=max_row)

#adding data
barchart.add_data(data, titles_from_data=True)
#adding categories
barchart.set_categories(catogeries)

#indicate where you want to add chart
sheet.add_chart(barchart,"H8")

barchart.title="spending by gender"
barchart.style=5
wb.save('barchart.xlsx')



