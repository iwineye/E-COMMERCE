from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
#import os
#import sys

wb=load_workbook('barchart.xlsx')
sheet=wb['spending_gender']

min_column=wb.active.min_column
max_column=wb.active.max_column
min_row=wb.active.min_row
max_row=wb.active.max_row

#sheet['B5']='=SUM(B2:B4)'
#sheet['B5'].style='Currency'
for i in range(min_column+1, max_column+2):
    letter=get_column_letter(i)
    sheet[f'{letter}{max_row+1}']=f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style='Currency'


wb.save('operations.xlsx')